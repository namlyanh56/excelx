#!/usr/bin/env python3
"""
Inventori Module
- Menu: Inventori
- Submenu: Start, Input, Info
- Penyimpanan data sementara per user (in-memory)
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from enum import IntEnum
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from telegram import (
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    InputFile,
    KeyboardButton,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    Update,
)
from telegram.constants import ParseMode
from telegram.ext import (
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)

INVENTORY_MENU_LABEL = "📦 Stored XIE"
INVENTORY_SUBMENU_START = "🚀 Start"
INVENTORY_SUBMENU_INPUT = "➕ Input"
INVENTORY_SUBMENU_INFO = "📊 Info"
INVENTORY_SUBMENU_BACK = "🔙 Kembali"

# Telegram UI Message Effects IDs
EFFECT_FIRE = "5104841245755180586"
EFFECT_TADA = "5046509860389126442"

PASSWORD_REGEX = re.compile(r"^[^\s]{6,64}$")
UID_REGEX = re.compile(r"^\d{8,20}$")
FILENAME_REGEX = re.compile(r"^[A-Za-z0-9_-]{1,50}$")
COOKIE_UID_REGEX = re.compile(r"(?:^|;)\s*c_user=(\d+)")
COOKIE_XS_REGEX = re.compile(r"(?:^|;)\s*xs=")


class InventoryStates(IntEnum):
    ASK_COOKIE = 201
    ASK_PASSWORD = 202
    ASK_FILENAME = 203


@dataclass
class InventoryEntry:
    uid: str
    password: str
    cookie: str
    created_at: str


@dataclass
class InventoryMeta:
    last_uid: str = ""
    last_password_empty: bool = False
    last_input_at: str = ""


# In-memory store (per user)
_INVENTORY_STORE: Dict[int, List[InventoryEntry]] = {}
_INVENTORY_META: Dict[int, InventoryMeta] = {}

# guard_access will be injected from main app
_GUARD_ACCESS = None


def configure_guard(guard_access):
    global _GUARD_ACCESS
    _GUARD_ACCESS = guard_access


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def _user_id(update: Update) -> Optional[int]:
    return update.effective_user.id if update.effective_user else None


def _inventory_menu_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(INVENTORY_SUBMENU_INPUT), KeyboardButton(INVENTORY_SUBMENU_INFO)],
            [KeyboardButton(INVENTORY_SUBMENU_START), KeyboardButton(INVENTORY_SUBMENU_BACK)],
        ],
        resize_keyboard=True,
        one_time_keyboard=False,
    )


def _skip_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("⏭️ Skip Password", callback_data="inv_skip_password")]
    ])


def _inline_cancel_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("❌ Batal & Kembali", callback_data="inv_cancel_input")]
    ])


def _validate_cookie_minimal(cookie: str) -> tuple[bool, str]:
    if "c_user=" not in cookie:
        return False, "Cookie wajib mengandung <code>c_user=</code>."
    if not COOKIE_XS_REGEX.search(cookie):
        return False, "Cookie wajib mengandung <code>xs=</code>."
    return True, ""


def _extract_uid(cookie: str) -> Optional[str]:
    m = COOKIE_UID_REGEX.search(cookie)
    if not m:
        return None
    uid = m.group(1).strip()
    if not UID_REGEX.fullmatch(uid):
        return None
    return uid


def _validate_filename(raw: str) -> tuple[bool, str]:
    s = raw.strip()
    if not s:
        return True, ""
    if not FILENAME_REGEX.fullmatch(s):
        return (
            False,
            "❌ <b>Nama file tidak valid.</b>\n\n"
            "Gunakan hanya huruf, angka, underscore (_), dash (-), maksimal 50 karakter."
        )
    return True, ""


def _build_filename(raw: str) -> str:
    s = raw.strip()
    if not s:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"inventory_export_{ts}.xlsx"
    return f"{s}.xlsx"


def _store_entry(user_id: int, uid: str, password: str, cookie: str) -> None:
    entries = _INVENTORY_STORE.setdefault(user_id, [])
    entries.append(
        InventoryEntry(
            uid=uid,
            password=password,
            cookie=cookie,
            created_at=_utc_now_iso(),
        )
    )
    meta = _INVENTORY_META.setdefault(user_id, InventoryMeta())
    meta.last_uid = uid
    meta.last_password_empty = password == ""
    meta.last_input_at = _utc_now_iso()


async def _guard(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    if _GUARD_ACCESS is None:
        return True
    return await _GUARD_ACCESS(update, context)


# -----------------------------------------------------------------------------
# Menu Handler
# -----------------------------------------------------------------------------

async def inventory_menu_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not await _guard(update, context):
        return ConversationHandler.END

    await update.effective_message.reply_text(
        "📦 <b>Sistem Inventori Aktif</b>\n\n"
        "Area untuk mengumpulkan dan menyimpan sementara data akun Anda sebelum dicetak menjadi Excel.\n\n"
        "👉 <i>Pilih aksi yang ingin Anda lakukan:</i>",
        parse_mode=ParseMode.HTML,
        reply_markup=_inventory_menu_keyboard(),
        message_effect_id=EFFECT_FIRE
    )
    return ConversationHandler.END


# -----------------------------------------------------------------------------
# Input Flow
# -----------------------------------------------------------------------------

async def inventory_input_start_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not await _guard(update, context):
        return ConversationHandler.END

    context.user_data.pop("inv_pending", None)
    
    # Hapus ReplyKeyboard lama agar antarmuka terlihat bersih saat proses inline
    try:
        tmp_msg = await update.effective_message.reply_text(
            "🔄 <i>Menyiapkan UI...</i>",
            parse_mode=ParseMode.HTML,
            reply_markup=ReplyKeyboardRemove()
        )
        await tmp_msg.delete()
    except Exception:
        pass

    await update.effective_message.reply_text(
        "🧾 <b>Input Inventori [Langkah 1/2]</b>\n\n"
        "👉 Silakan kirimkan <b>Cookie Lengkap</b> dari akun Anda.",
        parse_mode=ParseMode.HTML,
        reply_markup=_inline_cancel_keyboard(),
    )
    return InventoryStates.ASK_COOKIE


async def inventory_cookie_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    raw = (update.effective_message.text or "").strip()
    
    # Deteksi fallback kembali
    if raw in {INVENTORY_SUBMENU_BACK, "Batal", "/start"}:
        await update.effective_message.reply_text(
            "❎ Input dibatalkan.",
            parse_mode=ParseMode.HTML,
            reply_markup=_inventory_menu_keyboard()
        )
        context.user_data.pop("inv_pending", None)
        return ConversationHandler.END

    ok, reason = _validate_cookie_minimal(raw)
    if not ok:
        await update.effective_message.reply_text(
            f"❌ <b>Cookie Ditolak.</b>\n💡 {reason}",
            parse_mode=ParseMode.HTML,
            reply_markup=_inline_cancel_keyboard(),
        )
        return InventoryStates.ASK_COOKIE

    uid = _extract_uid(raw)
    if not uid:
        await update.effective_message.reply_text(
            "❌ <b>Gagal Mendeteksi UID.</b>\n"
            "Pastikan cookie memiliki <code>c_user=</code> yang terformat dengan benar.",
            parse_mode=ParseMode.HTML,
            reply_markup=_inline_cancel_keyboard(),
        )
        return InventoryStates.ASK_COOKIE

    context.user_data["inv_pending"] = {"uid": uid, "cookie": raw}
    await update.effective_message.reply_text(
        f"✅ <b>Cookie Valid!</b> (UID: <code>{uid}</code>)\n\n"
        "🔐 <b>Input Inventori [Langkah 2/2]</b>\n"
        "👉 Silakan masukkan <b>Password</b> untuk akun ini.\n"
        "<i>* Anda dapat menekan tombol Skip jika tidak ingin menyertakan password.</i>",
        parse_mode=ParseMode.HTML,
        reply_markup=_skip_keyboard(),
    )
    return InventoryStates.ASK_PASSWORD


async def inventory_password_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    raw = (update.effective_message.text or "").strip()
    
    if raw in {INVENTORY_SUBMENU_BACK, "Batal", "/start"}:
        await update.effective_message.reply_text(
            "❎ Input dibatalkan.",
            parse_mode=ParseMode.HTML,
            reply_markup=_inventory_menu_keyboard()
        )
        context.user_data.pop("inv_pending", None)
        return ConversationHandler.END

    if not PASSWORD_REGEX.fullmatch(raw):
        await update.effective_message.reply_text(
            "❌ <b>Password Tidak Valid.</b>\n"
            "📌 Syarat: 6–64 karakter dan tidak boleh ada spasi.\n"
            "<i>(Gunakan tombol <b>Skip Password</b> di atas jika kosong)</i>",
            parse_mode=ParseMode.HTML,
        )
        return InventoryStates.ASK_PASSWORD

    pending = context.user_data.get("inv_pending", {})
    uid = pending.get("uid", "")
    cookie = pending.get("cookie", "")
    if not uid or not cookie:
        await update.effective_message.reply_text(
            "❌ <b>Sesi Pending Hilang.</b> Silakan ulangi input dari awal.",
            parse_mode=ParseMode.HTML,
            reply_markup=_inventory_menu_keyboard(),
        )
        return ConversationHandler.END

    user_id = _user_id(update)
    if user_id is None:
        return ConversationHandler.END

    _store_entry(user_id, uid, raw, cookie)

    await update.effective_message.reply_text(
        "✨ <b>Satu Entri Berhasil Ditambahkan ke Inventori!</b>\n\n"
        f"👤 <b>UID:</b> <code>{uid}</code>\n"
        f"🔑 <b>Password:</b> <code>{raw}</code>\n"
        f"🕒 <b>Waktu:</b> <code>{_utc_now_iso()}</code>",
        parse_mode=ParseMode.HTML,
        reply_markup=_inventory_menu_keyboard(),
        message_effect_id=EFFECT_TADA
    )
    context.user_data.pop("inv_pending", None)
    return ConversationHandler.END


async def inventory_password_skip_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer("Memproses penyimpanan tanpa password...", show_alert=False)

    pending = context.user_data.get("inv_pending", {})
    uid = pending.get("uid", "")
    cookie = pending.get("cookie", "")
    if not uid or not cookie:
        await query.edit_message_text(
            "❌ <b>Sesi Pending Hilang.</b> Silakan ulangi input dari awal.",
            parse_mode=ParseMode.HTML,
        )
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="🔙 <b>Kembali ke Menu Inventori.</b>",
            parse_mode=ParseMode.HTML,
            reply_markup=_inventory_menu_keyboard()
        )
        return ConversationHandler.END

    user_id = _user_id(update)
    if user_id is None:
        return ConversationHandler.END

    _store_entry(user_id, uid, "", cookie)

    await query.edit_message_text(
        "✨ <b>Satu Entri Berhasil Ditambahkan ke Inventori!</b>\n\n"
        f"👤 <b>UID:</b> <code>{uid}</code>\n"
        f"🔑 <b>Password:</b> <i>(Dikosongkan)</i>\n"
        f"🕒 <b>Waktu:</b> <code>{_utc_now_iso()}</code>",
        parse_mode=ParseMode.HTML,
    )
    
    await context.bot.send_message(
        chat_id=query.message.chat_id,
        text="✨ <b>Data telah diamankan di Inventori.</b>",
        parse_mode=ParseMode.HTML,
        reply_markup=_inventory_menu_keyboard(),
        message_effect_id=EFFECT_TADA
    )

    context.user_data.pop("inv_pending", None)
    return ConversationHandler.END


async def inventory_cancel_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer("Aksi dibatalkan.", show_alert=False)
    
    context.user_data.pop("inv_pending", None)
    
    await query.edit_message_text(
        "❎ <i>Anda membatalkan aksi pada sesi ini. Proses dihentikan.</i>",
        parse_mode=ParseMode.HTML
    )
    
    await context.bot.send_message(
        chat_id=query.message.chat_id,
        text="🔙 <b>Kembali ke Menu Inventori.</b>",
        parse_mode=ParseMode.HTML,
        reply_markup=_inventory_menu_keyboard()
    )
    return ConversationHandler.END


# -----------------------------------------------------------------------------
# Start Flow (Ask Filename & Generate XLSX)
# -----------------------------------------------------------------------------

def _build_inventory_xlsx(entries: List[InventoryEntry]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "INVENTORY"

    headers = ["UID", "PASSWORD", "COOKIE"]
    ws.append(headers)

    for e in entries:
        ws.append([e.uid, e.password, e.cookie])

    header_font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    medium_side = Side(style="medium", color="000000")
    all_border = Border(left=medium_side, right=medium_side, top=medium_side, bottom=medium_side)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = all_border

    max_row = ws.max_row
    for r in range(2, max_row + 1):
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            cell.border = all_border
            cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=(c == 3),
            )

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 80
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{max_row}"
    ws.row_dimensions[1].height = 24

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def inventory_start_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not await _guard(update, context):
        return ConversationHandler.END

    user_id = _user_id(update)
    if user_id is None:
        return ConversationHandler.END

    entries = _INVENTORY_STORE.get(user_id, [])
    if not entries:
        await update.effective_message.reply_text(
            "⚠️ <b>Inventori Anda Masih Kosong!</b>\n"
            "Silakan tambahkan minimal 1 akun melalui menu <b>➕ Input</b> sebelum membuat file Excel.",
            parse_mode=ParseMode.HTML,
            reply_markup=_inventory_menu_keyboard(),
        )
        return ConversationHandler.END

    # Persiapkan UI inline
    try:
        tmp_msg = await update.effective_message.reply_text(
            "🔄 <i>Menyiapkan UI...</i>",
            parse_mode=ParseMode.HTML,
            reply_markup=ReplyKeyboardRemove()
        )
        await tmp_msg.delete()
    except Exception:
        pass

    await update.effective_message.reply_text(
        "📝 <b>Pembuatan File Inventori</b>\n\n"
        "👉 <b>Masukkan Nama File Excel</b> (tanpa .xlsx)\n"
        "<i>* Kosongkan pesan (atau ketik bebas) jika ingin menggunakan nama waktu otomatis.</i>",
        parse_mode=ParseMode.HTML,
        reply_markup=_inline_cancel_keyboard(),
    )
    return InventoryStates.ASK_FILENAME


async def inventory_filename_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    raw = (update.effective_message.text or "").strip()

    if raw in {INVENTORY_SUBMENU_BACK, "Batal", "/start"}:
        await update.effective_message.reply_text(
            "❎ Proses pembuatan file dibatalkan.",
            parse_mode=ParseMode.HTML,
            reply_markup=_inventory_menu_keyboard()
        )
        return ConversationHandler.END

    ok, err = _validate_filename(raw)
    if not ok:
        await update.effective_message.reply_text(
            err,
            parse_mode=ParseMode.HTML,
            reply_markup=_inline_cancel_keyboard(),
        )
        return InventoryStates.ASK_FILENAME

    user_id = _user_id(update)
    entries = _INVENTORY_STORE.get(user_id, [])
    
    if not entries:
        await update.effective_message.reply_text(
            "⚠️ Data inventori kosong secara tiba-tiba.",
            parse_mode=ParseMode.HTML,
            reply_markup=_inventory_menu_keyboard(),
        )
        return ConversationHandler.END

    filename = _build_filename(raw)
    buffer = _build_inventory_xlsx(entries)
    
    await update.effective_chat.send_message(
        text=f"✨ <b>Sukses Membuat Dokumen!</b>\nTotal <b>{len(entries)}</b> akun telah dirender ke dalam Excel dan <b>data inventori Anda telah di-reset (dikosongkan) otomatis</b>. Silakan unduh file Anda di bawah ini.",
        parse_mode=ParseMode.HTML,
        message_effect_id=EFFECT_TADA
    )

    await update.effective_chat.send_document(
        document=InputFile(buffer, filename=filename),
        reply_markup=_inventory_menu_keyboard(),
    )
    
    # Reset/clear the in-memory store for the user after generating the document
    if user_id in _INVENTORY_STORE:
        del _INVENTORY_STORE[user_id]
    if user_id in _INVENTORY_META:
        del _INVENTORY_META[user_id]

    return ConversationHandler.END


# -----------------------------------------------------------------------------
# Info Flow
# -----------------------------------------------------------------------------

async def inventory_info_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not await _guard(update, context):
        return ConversationHandler.END

    user_id = _user_id(update)
    if user_id is None:
        return ConversationHandler.END

    entries = _INVENTORY_STORE.get(user_id, [])
    meta = _INVENTORY_META.get(user_id, InventoryMeta())

    if not entries:
        await update.effective_message.reply_text(
            "ℹ️ <b>Status Inventori Kosong</b>\nBelum ada akun yang Anda masukkan ke dalam sistem saat ini.",
            parse_mode=ParseMode.HTML,
            reply_markup=_inventory_menu_keyboard(),
        )
        return ConversationHandler.END

    password_status = "❌ Tidak Ada" if meta.last_password_empty else "✅ Ada"
    await update.effective_message.reply_text(
        "📊 <b>Ringkasan Inventori Anda</b>\n\n"
        f"📦 <b>Total Akun Tersimpan:</b> <code>{len(entries)}</code> akun\n"
        f"👤 <b>UID Input Terakhir:</b> <code>{meta.last_uid}</code>\n"
        f"🔑 <b>Status Password Terakhir:</b> {password_status}\n"
        f"🕒 <b>Waktu Input Terakhir:</b> <code>{meta.last_input_at}</code>\n",
        parse_mode=ParseMode.HTML,
        reply_markup=_inventory_menu_keyboard(),
    )
    return ConversationHandler.END


# -----------------------------------------------------------------------------
# Registration Helper
# -----------------------------------------------------------------------------

def register_inventory_handlers(app, guard_access) -> None:
    configure_guard(guard_access)

    # Menu entry
    app.add_handler(CommandHandler("inventori", inventory_menu_handler))
    app.add_handler(MessageHandler(filters.Regex(f"^{re.escape(INVENTORY_MENU_LABEL)}$"), inventory_menu_handler))

    # Info submenu
    app.add_handler(MessageHandler(filters.Regex(f"^{re.escape(INVENTORY_SUBMENU_INFO)}$"), inventory_info_handler))

    # Input conversation
    inv_conv = ConversationHandler(
        entry_points=[
            CommandHandler("inventori_input", inventory_input_start_handler),
            MessageHandler(filters.Regex(f"^{re.escape(INVENTORY_SUBMENU_INPUT)}$"), inventory_input_start_handler),
        ],
        states={
            InventoryStates.ASK_COOKIE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, inventory_cookie_handler)
            ],
            InventoryStates.ASK_PASSWORD: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, inventory_password_handler),
                # BUGFIX: Callback skip dipindahkan ke dalam state ASK_PASSWORD
                CallbackQueryHandler(inventory_password_skip_callback, pattern="^inv_skip_password$")
            ],
        },
        fallbacks=[
            CommandHandler("inventori", inventory_menu_handler),
            CallbackQueryHandler(inventory_cancel_callback, pattern="^inv_cancel_input$"),
        ],
        allow_reentry=True,
        name="inventori_conversation",
        persistent=False,
    )
    app.add_handler(inv_conv)

    # Generate/Start conversation
    gen_conv = ConversationHandler(
        entry_points=[
            MessageHandler(filters.Regex(f"^{re.escape(INVENTORY_SUBMENU_START)}$"), inventory_start_handler),
        ],
        states={
            InventoryStates.ASK_FILENAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, inventory_filename_handler)],
        },
        fallbacks=[
            CommandHandler("inventori", inventory_menu_handler),
            CallbackQueryHandler(inventory_cancel_callback, pattern="^inv_cancel_input$"),
        ],
        allow_reentry=True,
        name="inventori_gen_conversation",
        persistent=False,
    )
    app.add_handler(gen_conv)
