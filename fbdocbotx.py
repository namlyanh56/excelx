#!/usr/bin/env python3
"""
Telegram Bot: Facebook Account XLSX Generator
- Python 3.11+
- python-telegram-bot v20+
- openpyxl
- Modular handler architecture in a single file
- Strict validation (rule-based)
- ConversationHandler for interactive flows
- UI/UX Revamped with Inline Cancel & Message Effects
"""

from __future__ import annotations

import io
import json
import logging
import os
import re
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from enum import IntEnum
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from telegram import (
    InputFile,
    KeyboardButton,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Update,
)
from telegram.constants import ParseMode
from telegram.ext import (
    Application,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    CallbackQueryHandler,
    MessageHandler,
    filters,
)

import inventori

# -----------------------------------------------------------------------------
# Environment & Logging
# -----------------------------------------------------------------------------

load_dotenv()

LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
logging.basicConfig(
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    level=getattr(logging, LOG_LEVEL, logging.INFO),
)
logger = logging.getLogger("fbdocbotx")

# -----------------------------------------------------------------------------
# Constants & Message Effects
# -----------------------------------------------------------------------------

MAIN_MENU_START = "🚀 Mulai Ulang"
MAIN_MENU_CREATE_DOC = "📝 Buat Dokumen Excel"
MAIN_MENU_HELP = "ℹ️ Pusat Bantuan"
MAIN_MENU_ADMIN = "🛡️ Admin Panel"
MAIN_MENU_INVENTORY = inventori.INVENTORY_MENU_LABEL

SUBMENU_MANUAL = "⌨️ Input Manual"
SUBMENU_INSTANT = "⚡ Input Instan"
SUBMENU_BACK = "🔙 Kembali"
SUBMENU_CANCEL = "❌ Batal"  # Masih dipertahankan untuk fallback text

ADMIN_MENU_WHITELIST = "✅ Kelola Whitelist"
ADMIN_MENU_DURATION = "⏳ Atur Durasi Akses"
ADMIN_MENU_BLOCK = "⛔ Blokir User"
ADMIN_MENU_BROADCAST = "📣 Broadcast"
ADMIN_MENU_STATS = "📊 Statistik"
ADMIN_MENU_EXTRA = "🧩 Admin Tambahan"

# Telegram UI Message Effects IDs
EFFECT_FIRE = "5104841245755180586"
EFFECT_TADA = "5046509860389126442"

UID_REGEX = re.compile(r"^[0-9]{8,20}$")
PASSWORD_REGEX = re.compile(r"^[^\s]{6,64}$")
FILENAME_REGEX = re.compile(r"^[A-Za-z0-9_-]{1,50}$")

# Delimiter split: comma, whitespace (space/tab/newline), including multiple
SPLIT_REGEX = re.compile(r"[,\s]+")

# Strict cookie key=value; validator (semicolon optional at end)
COOKIE_FORMAT_REGEX = re.compile(
    r"^\s*[A-Za-z0-9_]+=[^;=\n\r]+(?:;\s*[A-Za-z0-9_]+=[^;=\n\r]+)*;?\s*$"
)

DATA_STORE_FILE = Path("bot_data.json")


class States(IntEnum):
    ASK_UID = 1
    ASK_PASSWORD = 2
    ASK_COOKIE = 3
    ASK_FILENAME_MANUAL = 4
    ASK_INSTANT_PAYLOAD = 5
    ASK_FILENAME_INSTANT = 6


class AdminStates(IntEnum):
    MENU = 101
    WHITELIST_INPUT = 102
    DURATION_INPUT = 103
    BLOCK_INPUT = 104
    BROADCAST_INPUT = 105


@dataclass
class ParsedInput:
    uids: List[str]
    passwords: List[str]
    cookies: List[str]


# -----------------------------------------------------------------------------
# Config / Admin helpers
# -----------------------------------------------------------------------------

def parse_admin_ids() -> set[int]:
    raw = os.getenv("ADMIN_IDS", "").strip()
    ids: set[int] = set()
    if not raw:
        return ids
    for x in raw.split(","):
        x = x.strip()
        if x.isdigit():
            ids.add(int(x))
    return ids


ADMIN_IDS = parse_admin_ids()


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def parse_utc_iso(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except Exception:
        return None


def to_utc_iso(dt: datetime) -> str:
    return dt.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")


def load_store() -> dict:
    if not DATA_STORE_FILE.exists():
        return {
            "users": {},
            "blocked": [],
            "stats": {
                "total_docs_created": 0,
                "total_messages_processed": 0,
                "last_broadcast_at": None,
            },
        }
    try:
        return json.loads(DATA_STORE_FILE.read_text(encoding="utf-8"))
    except Exception:
        logger.exception("Failed to load store; fallback to default")
        return {
            "users": {},
            "blocked": [],
            "stats": {
                "total_docs_created": 0,
                "total_messages_processed": 0,
                "last_broadcast_at": None,
            },
        }


def save_store(store: dict) -> None:
    DATA_STORE_FILE.write_text(
        json.dumps(store, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def ensure_user_record(store: dict, user_id: int) -> dict:
    users: Dict[str, dict] = store.setdefault("users", {})
    key = str(user_id)
    if key not in users:
        users[key] = {
            "whitelisted": False,
            "access_expires_at": None,
            "created_docs": 0,
            "last_seen_at": to_utc_iso(utc_now()),
        }
    return users[key]


def user_is_admin(user_id: int | None) -> bool:
    return bool(user_id and user_id in ADMIN_IDS)


def user_is_blocked(store: dict, user_id: int | None) -> bool:
    if not user_id:
        return False
    return user_id in set(store.get("blocked", []))


def user_has_access(store: dict, user_id: int | None) -> bool:
    if user_is_admin(user_id):
        return True
    if not user_id:
        return False
    if user_is_blocked(store, user_id):
        return False

    user = ensure_user_record(store, user_id)
    if not user.get("whitelisted", False):
        return False

    exp = parse_utc_iso(user.get("access_expires_at"))
    if exp is None:
        return True
    return utc_now() <= exp


def touch_user(store: dict, user_id: int | None) -> None:
    if not user_id:
        return
    u = ensure_user_record(store, user_id)
    u["last_seen_at"] = to_utc_iso(utc_now())


# -----------------------------------------------------------------------------
# UI / Keyboard Builders
# -----------------------------------------------------------------------------

def main_menu_keyboard(is_admin: bool = False) -> ReplyKeyboardMarkup:
    rows = [
        [KeyboardButton(MAIN_MENU_INVENTORY)],  # Dipindahkan ke paling atas
        [KeyboardButton(MAIN_MENU_CREATE_DOC)],
        [KeyboardButton(MAIN_MENU_START), KeyboardButton(MAIN_MENU_HELP)],
    ]
    if is_admin:
        rows.append([KeyboardButton(MAIN_MENU_ADMIN)])

    return ReplyKeyboardMarkup(
        keyboard=rows,
        resize_keyboard=True,
        one_time_keyboard=False,
    )


def create_doc_submenu_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(SUBMENU_MANUAL), KeyboardButton(SUBMENU_INSTANT)],
            [KeyboardButton(SUBMENU_BACK)],
        ],
        resize_keyboard=True,
        one_time_keyboard=False,
    )


def inline_cancel_keyboard() -> InlineKeyboardMarkup:
    """Mengembalikan keyboard inline untuk membatalkan proses yang sedang berjalan."""
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("❌ Batal & Kembali", callback_data="cancel_input")]
    ])


def admin_menu_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(ADMIN_MENU_WHITELIST), KeyboardButton(ADMIN_MENU_DURATION)],
            [KeyboardButton(ADMIN_MENU_BLOCK), KeyboardButton(ADMIN_MENU_BROADCAST)],
            [KeyboardButton(ADMIN_MENU_STATS), KeyboardButton(ADMIN_MENU_EXTRA)],
            [KeyboardButton(SUBMENU_BACK)],
        ],
        resize_keyboard=True,
        one_time_keyboard=False,
    )


# -----------------------------------------------------------------------------
# Parsing & Validation Utilities
# -----------------------------------------------------------------------------

def split_tokens(text: str) -> List[str]:
    return [part.strip() for part in SPLIT_REGEX.split(text.strip()) if part.strip()]


def has_delimiter(raw: str) -> bool:
    return bool(re.search(r"[,\s]", raw))


def validate_uids(uids: Sequence[str]) -> Tuple[bool, str]:
    if not uids:
        return False, "❌ <b>Oops! UID kosong.</b>\nSilakan masukkan minimal 1 UID."
    for i, uid in enumerate(uids, start=1):
        if not UID_REGEX.fullmatch(uid):
            return (
                False,
                f"❌ <b>UID ke-{i} tidak valid:</b> <code>{uid}</code>\n"
                "📌 <i>Syarat: Hanya digit, panjang 8–20 karakter.</i>",
            )
    return True, ""


def validate_passwords(passwords: Sequence[str]) -> Tuple[bool, str]:
    if not passwords:
        return False, "❌ <b>Oops! Password kosong.</b>\nSilakan masukkan minimal 1 password."
    for i, pwd in enumerate(passwords, start=1):
        if not PASSWORD_REGEX.fullmatch(pwd):
            return (
                False,
                f"❌ <b>Password ke-{i} tidak valid.</b>\n"
                "📌 <i>Syarat: 6–64 karakter dan tidak boleh mengandung spasi.</i>",
            )
    return True, ""


def validate_cookie(cookie: str) -> Tuple[bool, str]:
    c = cookie.strip()
    if not c:
        return False, "Cookie tidak boleh kosong."
    if len(c) < 20:
        return False, "Cookie minimal 20 karakter."
    if "c_user=" not in c or "xs=" not in c:
        return False, "Cookie wajib mengandung <code>c_user=</code> dan <code>xs=</code>."
    if not COOKIE_FORMAT_REGEX.fullmatch(c):
        return False, "Format harus <code>key=value;key=value;</code>."
    return True, ""


def validate_cookies(cookies: Sequence[str]) -> Tuple[bool, str]:
    if not cookies:
        return False, "❌ <b>Oops! Cookie kosong.</b>\nSilakan masukkan minimal 1 cookie."
    for i, ck in enumerate(cookies, start=1):
        ok, reason = validate_cookie(ck)
        if not ok:
            return False, f"❌ <b>Cookie ke-{i} tidak valid.</b>\n💡 <i>Alasan: {reason}</i>"
    return True, ""


def parse_instant_message(text: str) -> Tuple[bool, str, ParsedInput | None]:
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if len(lines) < 3:
        return (
            False,
            "❌ <b>Format Input Instan minimal 3 baris:</b>\n\n"
            "1️⃣ Baris 1: UID\n2️⃣ Baris 2: PASSWORD\n3️⃣ Baris 3: COOKIE",
            None,
        )

    uid_line, pwd_line, cookie_line = lines[0], lines[1], lines[2]

    uids = split_tokens(uid_line)
    passwords = split_tokens(pwd_line)
    cookies = split_tokens(cookie_line)

    ok, err = validate_uids(uids)
    if not ok:
        return False, err, None

    ok, err = validate_passwords(passwords)
    if not ok:
        return False, err, None

    ok, err = validate_cookies(cookies)
    if not ok:
        return False, err, None

    if not (len(uids) == len(passwords) == len(cookies)):
        return (
            False,
            "❌ <b>Jumlah data tidak seimbang!</b>\n\n"
            f"📊 UID: {len(uids)}\n🔑 PASSWORD: {len(passwords)}\n🍪 COOKIE: {len(cookies)}",
            None,
        )

    return True, "", ParsedInput(uids=uids, passwords=passwords, cookies=cookies)


def validate_filename_no_ext(raw: str) -> Tuple[bool, str]:
    s = raw.strip()
    if not s:
        return True, ""  # empty allowed => default timestamp name
    if not FILENAME_REGEX.fullmatch(s):
        return (
            False,
            "❌ <b>Nama file tidak valid.</b>\n\n"
            "Gunakan hanya huruf, angka, underscore (_), dash (-), maksimal 50 karakter.",
        )
    return True, ""


def build_filename(raw: str) -> str:
    s = raw.strip()
    if not s:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"facebook_accounts_{ts}.xlsx"
    return f"{s}.xlsx"


# -----------------------------------------------------------------------------
# XLSX Generator
# -----------------------------------------------------------------------------

def build_xlsx_file(data: ParsedInput) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "DATA"

    headers = ["UID", "PASSWORD", "COOKIE"]
    ws.append(headers)

    for uid, pwd, cookie in zip(data.uids, data.passwords, data.cookies):
        ws.append([uid, pwd, cookie])

    header_font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    medium_side = Side(style="medium", color="000000")
    all_border = Border(left=medium_side, right=medium_side, top=medium_side, bottom=medium_side)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = all_border

    max_row = ws.max_row
    for r in range(2, max_row + 1):
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            cell.border = all_border
            cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=(c == 3),
            )

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 80
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{max_row}"
    ws.row_dimensions[1].height = 24

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# -----------------------------------------------------------------------------
# Response Text
# -----------------------------------------------------------------------------

HELP_TEXT = (
    "📚 <b>PUSAT BANTUAN FBDOCBOT</b>\n\n"
    "🎯 <b>Menu Utama</b>\n"
    "• 🚀 <b>Mulai Ulang:</b> Membersihkan sesi dan memuat ulang antarmuka bot.\n"
    "• 📝 <b>Buat Dokumen:</b> Mengakses panel pembuatan file Excel.\n"
    "• ℹ️ <b>Bantuan:</b> Menampilkan panduan ringkas penggunaan bot ini.\n\n"
    "⌨️ <b>Input Manual</b>\n"
    "Mode bertahap untuk memasukkan data baris per baris:\n"
    "1️⃣ UID\n2️⃣ PASSWORD\n3️⃣ COOKIE\n4️⃣ Nama file\n\n"
    "⚡ <b>Input Instan</b>\n"
    "Kirim sekaligus dalam 1 pesan terstruktur (minimal 3 baris):\n"
    "Baris 1: Kumpulan UID\nBaris 2: Kumpulan PASSWORD\nBaris 3: Kumpulan COOKIE\n\n"
    "💡 <b>Aturan Validasi Sistem:</b>\n"
    "• UID: Wajib 8-20 digit angka murni\n"
    "• PASSWORD: 6-64 karakter, tanpa spasi\n"
    "• COOKIE: Wajib mengandung elemen <code>c_user=</code> dan <code>xs=</code>\n"
    "• Nama File: Karakter standar [A-Za-z0-9_-], max 50"
)


# -----------------------------------------------------------------------------
# Session / Access Helpers
# -----------------------------------------------------------------------------

def hard_reset_user_session(context: ContextTypes.DEFAULT_TYPE) -> None:
    context.user_data.clear()


def current_user_id(update: Update) -> int | None:
    return update.effective_user.id if update.effective_user else None


def is_control_reset_text(text: str) -> bool:
    return text in {MAIN_MENU_START, SUBMENU_CANCEL, "Start", "Batal", "/start"}


async def force_back_to_main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE, text: str) -> int:
    hard_reset_user_session(context)
    uid = current_user_id(update)
    is_admin = user_is_admin(uid)
    await update.effective_message.reply_text(
        text,
        parse_mode=ParseMode.HTML,
        reply_markup=main_menu_keyboard(is_admin=is_admin),
    )
    return ConversationHandler.END


async def guard_access(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    store = load_store()
    uid = current_user_id(update)
    if uid is None:
        return False

    touch_user(store, uid)
    store["stats"]["total_messages_processed"] = store["stats"].get("total_messages_processed", 0) + 1
    save_store(store)

    if user_is_blocked(store, uid):
        await update.effective_message.reply_text(
            "⛔ <b>Akses Anda sedang diblokir.</b>\nHubungi administrator bot jika ini sebuah kesalahan.",
            parse_mode=ParseMode.HTML,
            reply_markup=main_menu_keyboard(is_admin=user_is_admin(uid)),
        )
        return False

    if not user_has_access(store, uid):
        await update.effective_message.reply_text(
            "🔒 <b>Akses Terbatas</b>\nAkun Anda belum masuk dalam daftar putih (Whitelist).\nSilakan hubungi administrator untuk meminta akses.",
            parse_mode=ParseMode.HTML,
            reply_markup=main_menu_keyboard(is_admin=user_is_admin(uid)),
        )
        return False

    return True


async def clear_keyboard_ui(update: Update) -> None:
    """Helper untuk menghapus ReplyKeyboard lama sebelum mengirim tombol Inline."""
    try:
        tmp_msg = await update.effective_message.reply_text(
            "🔄 <i>Menyiapkan UI...</i>",
            parse_mode=ParseMode.HTML,
            reply_markup=ReplyKeyboardRemove()
        )
        await tmp_msg.delete()
    except Exception as e:
        logger.debug(f"Clear keyboard UI skipped: {e}")


# -----------------------------------------------------------------------------
# Send Result
# -----------------------------------------------------------------------------

async def send_xlsx_result(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
    data: ParsedInput,
    filename: str,
) -> None:
    try:
        xlsx_buffer = build_xlsx_file(data)

        # Pisahkan notifikasi text agar message effect (TADA 🎉) berjalan optimal
        await update.effective_chat.send_message(
            text="✨ <b>Dokumen berhasil digenerasi dengan sempurna!</b>\nSilakan unduh file Excel Anda di bawah ini. ✅",
            parse_mode=ParseMode.HTML,
            message_effect_id=EFFECT_TADA
        )

        await update.effective_chat.send_document(
            document=InputFile(xlsx_buffer, filename=filename),
        )

        store = load_store()
        uid = current_user_id(update)
        if uid:
            u = ensure_user_record(store, uid)
            u["created_docs"] = int(u.get("created_docs", 0)) + 1
        store["stats"]["total_docs_created"] = store["stats"].get("total_docs_created", 0) + 1
        save_store(store)

    except Exception:
        logger.exception("Failed to generate/send XLSX")
        await update.effective_message.reply_text(
            "❌ <b>Terjadi kesalahan sistem internal</b> saat merender file XLSX. Mohon coba lagi.",
            parse_mode=ParseMode.HTML,
            reply_markup=inline_cancel_keyboard(),
        )


# -----------------------------------------------------------------------------
# Core Handlers
# -----------------------------------------------------------------------------

async def start_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    hard_reset_user_session(context)
    uid = current_user_id(update)
    is_admin = user_is_admin(uid)
    await update.effective_message.reply_text(
        "✨ <b>Selamat Datang di FBDocBot!</b>\n\n"
        "Asisten pintar Anda untuk menyusun dan mengelola dokumen Excel secara otomatis.\n\n"
        "👉 <i>Silakan pilih opsi dari menu di bawah untuk memulai.</i>",
        parse_mode=ParseMode.HTML,
        reply_markup=main_menu_keyboard(is_admin=is_admin),
        message_effect_id=EFFECT_FIRE
    )
    return ConversationHandler.END


async def help_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    uid = current_user_id(update)
    await update.effective_message.reply_text(
        HELP_TEXT,
        parse_mode=ParseMode.HTML,
        reply_markup=main_menu_keyboard(is_admin=user_is_admin(uid)),
    )
    return ConversationHandler.END


async def cancel_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    return await force_back_to_main_menu(
        update,
        context,
        "❎ <b>Sesi Dibatalkan.</b>\nSemua state sementara telah dihapus bersih. Kembali ke Menu Utama.",
    )


async def cancel_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handler spesifik untuk menangani tombol Batal jenis Inline."""
    query = update.callback_query
    await query.answer("Sesi aktif dibatalkan.", show_alert=False)
    hard_reset_user_session(context)

    # Edit pesan yang tadinya memiliki tombol Inline agar rapi
    await query.edit_message_text(
        "❎ <i>Anda membatalkan input data pada sesi ini. Proses dihentikan.</i>",
        parse_mode=ParseMode.HTML
    )

    uid = current_user_id(update)
    is_admin = user_is_admin(uid)
    # Kembalikan main menu
    await context.bot.send_message(
        chat_id=query.message.chat_id,
        text="🔙 <b>Kembali ke Menu Utama.</b>",
        parse_mode=ParseMode.HTML,
        reply_markup=main_menu_keyboard(is_admin=is_admin)
    )
    return ConversationHandler.END


async def menu_create_doc_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not await guard_access(update, context):
        return ConversationHandler.END

    hard_reset_user_session(context)
    await update.effective_message.reply_text(
        "🛠️ <b>Mode Pembuatan Dokumen</b>\n\n"
        "Pilih metode penyusunan data yang paling sesuai dengan kebutuhan Anda:",
        parse_mode=ParseMode.HTML,
        reply_markup=create_doc_submenu_keyboard(),
    )
    return ConversationHandler.END


# -----------------------------------------------------------------------------
# Manual Flow
# -----------------------------------------------------------------------------

async def manual_start_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not await guard_access(update, context):
        return ConversationHandler.END

    hard_reset_user_session(context)
    await clear_keyboard_ui(update)

    await update.effective_message.reply_text(
        "⌨️ <b>Input Manual [Langkah 1 / 4]</b>\n\n"
        "👉 <b>Masukkan daftar UID</b>\n"
        "<i>(Pisahkan antar data menggunakan spasi, koma, atau baris baru)</i>",
        parse_mode=ParseMode.HTML,
        reply_markup=inline_cancel_keyboard(),
    )
    return States.ASK_UID


async def ask_uid_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    raw = (update.effective_message.text or "").strip()
    if is_control_reset_text(raw):
        return await cancel_handler(update, context)

    if not has_delimiter(raw) and len(raw.split()) > 1:
        await update.effective_message.reply_text(
            "❌ <b>Format UID ditolak.</b> Harap pastikan setiap UID dipisahkan oleh spasi atau koma.",
            parse_mode=ParseMode.HTML,
            reply_markup=inline_cancel_keyboard(),
        )
        return States.ASK_UID

    uids = split_tokens(raw)
    ok, err = validate_uids(uids)
    if not ok:
        await update.effective_message.reply_text(
            err,
            parse_mode=ParseMode.HTML,
            reply_markup=inline_cancel_keyboard(),
        )
        return States.ASK_UID

    context.user_data["uids"] = uids
    await update.effective_message.reply_text(
        f"✅ <b>UID Valid!</b> Terdeteksi <b>{len(uids)}</b> entri data.\n\n"
        "🔐 <b>Input Manual [Langkah 2 / 4]</b>\n"
        "👉 <b>Masukkan daftar PASSWORD</b> dengan pemisah (spasi/koma/enter).",
        parse_mode=ParseMode.HTML,
        reply_markup=inline_cancel_keyboard(),
    )
    return States.ASK_PASSWORD


async def ask_password_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    raw = (update.effective_message.text or "").strip()
    if is_control_reset_text(raw):
        return await cancel_handler(update, context)

    passwords = split_tokens(raw)
    ok, err = validate_passwords(passwords)
    if not ok:
        await update.effective_message.reply_text(
            err,
            parse_mode=ParseMode.HTML,
            reply_markup=inline_cancel_keyboard(),
        )
        return States.ASK_PASSWORD

    uids = context.user_data.get("uids", [])
    if len(passwords) != len(uids):
        await update.effective_message.reply_text(
            f"❌ <b>Kuantitas Data Tidak Sinkron</b>\n"
            f"Total UID: <b>{len(uids)}</b>\n"
            f"Total PASSWORD: <b>{len(passwords)}</b>\n\n"
            "Harap perbaiki input Password Anda agar jumlahnya cocok.",
            parse_mode=ParseMode.HTML,
            reply_markup=inline_cancel_keyboard(),
        )
        return States.ASK_PASSWORD

    context.user_data["passwords"] = passwords
    await update.effective_message.reply_text(
        "✅ <b>Password Valid!</b>\n\n"
        "🍪 <b>Input Manual [Langkah 3 / 4]</b>\n"
        "👉 <b>Masukkan daftar COOKIE</b> dengan format yang tepat.",
        parse_mode=ParseMode.HTML,
        reply_markup=inline_cancel_keyboard(),
    )
    return States.ASK_COOKIE


async def ask_cookie_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    raw = (update.effective_message.text or "").strip()
    if is_control_reset_text(raw):
        return await cancel_handler(update, context)

    cookies = split_tokens(raw)
    ok, err = validate_cookies(cookies)
    if not ok:
        await update.effective_message.reply_text(
            err,
            parse_mode=ParseMode.HTML,
            reply_markup=inline_cancel_keyboard(),
        )
        return States.ASK_COOKIE

    uids = context.user_data.get("uids", [])
    passwords = context.user_data.get("passwords", [])
    if not (len(uids) == len(passwords) == len(cookies)):
        await update.effective_message.reply_text(
            "❌ <b>Kuantitas Data Tidak Seimbang</b>\n"
            "Jumlah UID, Password, dan Cookie harus presisi sama.",
            parse_mode=ParseMode.HTML,
            reply_markup=inline_cancel_keyboard(),
        )
        return States.ASK_COOKIE

    context.user_data["cookies"] = cookies
    await update.effective_message.reply_text(
        "📝 <b>Input Manual [Langkah 4 / 4]</b>\n\n"
        "👉 <b>Tentukan Nama File Excel</b> (tanpa .xlsx)\n"
        "<i>* Anda dapat mengosongkan/mengirim karakter apapun jika ingin sistem menggunakan penamaan waktu otomatis.</i>",
        parse_mode=ParseMode.HTML,
        reply_markup=inline_cancel_keyboard(),
    )
    return States.ASK_FILENAME_MANUAL


async def ask_filename_manual_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    raw = (update.effective_message.text or "").strip()
    if is_control_reset_text(raw):
        return await cancel_handler(update, context)

    ok, err = validate_filename_no_ext(raw)
    if not ok:
        await update.effective_message.reply_text(
            err,
            parse_mode=ParseMode.HTML,
            reply_markup=inline_cancel_keyboard(),
        )
        return States.ASK_FILENAME_MANUAL

    filename = build_filename(raw)
    parsed = ParsedInput(
        uids=context.user_data.get("uids", []),
        passwords=context.user_data.get("passwords", []),
        cookies=context.user_data.get("cookies", []),
    )

    progress_msg = await update.effective_message.reply_text(
        "⏳ <i>Memproses data yang terkumpul dan merender file Excel...</i>",
        parse_mode=ParseMode.HTML,
    )
    
    await send_xlsx_result(update, context, parsed, filename)
    await progress_msg.delete()

    return await force_back_to_main_menu(
        update,
        context,
        "🔙 Sesi Manual selesai. Anda telah kembali ke antarmuka utama.",
    )


# -----------------------------------------------------------------------------
# Instant Flow
# -----------------------------------------------------------------------------

async def instant_start_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not await guard_access(update, context):
        return ConversationHandler.END

    hard_reset_user_session(context)
    await clear_keyboard_ui(update)

    await update.effective_message.reply_text(
        "⚡ <b>Mode Input Instan Aktif</b>\n\n"
        "Kirim 1 pesan utuh berisi minimal 3 baris terpisah:\n"
        "<code>Baris 1</code>: Seluruh UID\n"
        "<code>Baris 2</code>: Seluruh PASSWORD\n"
        "<code>Baris 3</code>: Seluruh COOKIE\n\n"
        "👉 <i>Silakan masukkan payload instan Anda sekarang.</i>",
        parse_mode=ParseMode.HTML,
        reply_markup=inline_cancel_keyboard(),
    )
    return States.ASK_INSTANT_PAYLOAD


async def ask_instant_payload_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = (update.effective_message.text or "").strip()
    if is_control_reset_text(text):
        return await cancel_handler(update, context)

    ok, err, parsed = parse_instant_message(text)
    if not ok or parsed is None:
        await update.effective_message.reply_text(
            err,
            parse_mode=ParseMode.HTML,
            reply_markup=inline_cancel_keyboard(),
        )
        return States.ASK_INSTANT_PAYLOAD

    context.user_data["instant_parsed"] = {
        "uids": parsed.uids,
        "passwords": parsed.passwords,
        "cookies": parsed.cookies,
    }
    await update.effective_message.reply_text(
        "✅ <b>Data Instan Berhasil Divalidasi!</b>\n\n"
        "📝 Langkah Terakhir: <b>Masukkan nama file keluaran</b> (tanpa .xlsx).\n"
        "<i>* Cukup balas dengan karakter kosong untuk penamaan waktu otomatis.</i>",
        parse_mode=ParseMode.HTML,
        reply_markup=inline_cancel_keyboard(),
    )
    return States.ASK_FILENAME_INSTANT


async def ask_filename_instant_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    raw = (update.effective_message.text or "").strip()
    if is_control_reset_text(raw):
        return await cancel_handler(update, context)

    ok, err = validate_filename_no_ext(raw)
    if not ok:
        await update.effective_message.reply_text(
            err,
            parse_mode=ParseMode.HTML,
            reply_markup=inline_cancel_keyboard(),
        )
        return States.ASK_FILENAME_INSTANT

    filename = build_filename(raw)
    d = context.user_data.get("instant_parsed", {})
    parsed = ParsedInput(
        uids=d.get("uids", []),
        passwords=d.get("passwords", []),
        cookies=d.get("cookies", []),
    )

    progress_msg = await update.effective_message.reply_text(
        "⏳ <i>Memparsing payload instan dan menyusun matriks Excel...</i>",
        parse_mode=ParseMode.HTML,
    )
    
    await send_xlsx_result(update, context, parsed, filename)
    await progress_msg.delete()

    return await force_back_to_main_menu(
        update,
        context,
        "🔙 Sesi Instan selesai. Sistem dikembalikan ke posisi standby.",
    )


# -----------------------------------------------------------------------------
# Admin Handlers
# -----------------------------------------------------------------------------

async def admin_entry_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    uid = current_user_id(update)
    if not user_is_admin(uid):
        await update.effective_message.reply_text(
            "⛔ <b>Akses Ditolak Terotorisasi.</b>\nArea ini dikhususkan secara eksklusif untuk staf Administrator bot.",
            parse_mode=ParseMode.HTML,
            reply_markup=main_menu_keyboard(is_admin=False),
        )
        return ConversationHandler.END

    await update.effective_message.reply_text(
        "🛡️ <b>Masuk ke Panel Kontrol Administrator</b>\nSilakan tentukan aksi manajerial yang ingin Anda lakukan dari opsi di bawah:",
        parse_mode=ParseMode.HTML,
        reply_markup=admin_menu_keyboard(),
    )
    return AdminStates.MENU


async def admin_menu_router(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = (update.effective_message.text or "").strip()

    if text in {SUBMENU_CANCEL, MAIN_MENU_START}:
        return await cancel_handler(update, context)

    if text == SUBMENU_BACK:
        return await force_back_to_main_menu(update, context, "🔙 Anda telah keluar dari Panel Kontrol Admin.")

    if text == ADMIN_MENU_WHITELIST:
        await update.effective_message.reply_text(
            "📋 <b>Mode Pengelolaan Whitelist</b>\nKirim format eksekusi:\n"
            "• <code>allow [ID_PENGGUNA]</code> untuk mendaftarkan\n"
            "• <code>deny [ID_PENGGUNA]</code> untuk mencabut",
            parse_mode=ParseMode.HTML,
            reply_markup=admin_menu_keyboard(),
        )
        return AdminStates.WHITELIST_INPUT

    if text == ADMIN_MENU_DURATION:
        await update.effective_message.reply_text(
            "⏳ <b>Mode Pengaturan Durasi Expire</b>\nKirim format eksekusi:\n"
            "• <code>[ID_PENGGUNA] [JUMLAH_HARI]</code>\n"
            "<i>(Misal: 123456789 30 untuk akses aktif 30 hari)</i>",
            parse_mode=ParseMode.HTML,
            reply_markup=admin_menu_keyboard(),
        )
        return AdminStates.DURATION_INPUT

    if text == ADMIN_MENU_BLOCK:
        await update.effective_message.reply_text(
            "⛔ <b>Mode Restriksi Akses (Banned)</b>\nKirim format eksekusi:\n"
            "• <code>block [ID_PENGGUNA]</code> untuk memblokir permanen\n"
            "• <code>unblock [ID_PENGGUNA]</code> untuk melepaskan blokir",
            parse_mode=ParseMode.HTML,
            reply_markup=admin_menu_keyboard(),
        )
        return AdminStates.BLOCK_INPUT

    if text == ADMIN_MENU_BROADCAST:
        await update.effective_message.reply_text(
            "📣 <b>Mode Siaran Massa (Broadcast)</b>\n\nKirim pesan tekstual yang ingin Anda distribusikan ke seluruh pengguna di database:",
            parse_mode=ParseMode.HTML,
            reply_markup=admin_menu_keyboard(),
        )
        return AdminStates.BROADCAST_INPUT

    if text == ADMIN_MENU_STATS:
        store = load_store()
        users = store.get("users", {})
        blocked = store.get("blocked", [])
        docs = store.get("stats", {}).get("total_docs_created", 0)
        msgs = store.get("stats", {}).get("total_messages_processed", 0)

        await update.effective_message.reply_text(
            "📊 <b>Statistik Penggunaan FBDocBot</b>\n\n"
            f"👤 Total Populasi Database: <b>{len(users)}</b> pengguna\n"
            f"🚫 Pengguna dalam Daftar Hitam: <b>{len(blocked)}</b> pengguna\n"
            f"📄 Akumulasi Dokumen Digenerasi: <b>{docs}</b> file\n"
            f"💬 Keseluruhan Perintah Diproses: <b>{msgs}</b> pesan",
            parse_mode=ParseMode.HTML,
            reply_markup=admin_menu_keyboard(),
        )
        return AdminStates.MENU

    if text == ADMIN_MENU_EXTRA:
        await update.effective_message.reply_text(
            "🧩 <b>Modul Tambahan Administrator</b>\n\n"
            "Fitur ini direncanakan untuk pembaruan berikutnya:\n"
            "• 📥 Eksportasi Log Aktivitas Ringkas\n"
            "• 🔄 Pembersihan Hard Reset Metrik\n"
            "• 🕵️ Audit Jejak Perubahan Antar Admin",
            parse_mode=ParseMode.HTML,
            reply_markup=admin_menu_keyboard(),
        )
        return AdminStates.MENU

    await update.effective_message.reply_text(
        "Tolong gunakan hanya tata navigasi Menu Admin yang tertera.",
        parse_mode=ParseMode.HTML,
        reply_markup=admin_menu_keyboard(),
    )
    return AdminStates.MENU


async def admin_whitelist_input_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = (update.effective_message.text or "").strip()
    m = re.fullmatch(r"(allow|deny)\s+(\d+)", text, flags=re.IGNORECASE)
    if not m:
        await update.effective_message.reply_text(
            "❌ Sintaks gagal diurai. Mohon gunakan struktur valid: <code>allow 123456789</code> atau <code>deny 123456789</code>",
            parse_mode=ParseMode.HTML,
            reply_markup=admin_menu_keyboard(),
        )
        return AdminStates.WHITELIST_INPUT

    action = m.group(1).lower()
    target = int(m.group(2))

    store = load_store()
    u = ensure_user_record(store, target)
    u["whitelisted"] = action == "allow"
    save_store(store)

    await update.effective_message.reply_text(
        f"✅ Konfigurasi tersimpan: Whitelist untuk identitas <code>{target}</code> telah dideklarasikan menjadi <b>{u['whitelisted']}</b>.",
        parse_mode=ParseMode.HTML,
        reply_markup=admin_menu_keyboard(),
    )
    return AdminStates.MENU


async def admin_duration_input_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = (update.effective_message.text or "").strip()
    m = re.fullmatch(r"(\d+)\s+(\d+)", text)
    if not m:
        await update.effective_message.reply_text(
            "❌ Sintaks durasi tidak sah. Struktur standar yang diakui: <code>123456789 30</code>",
            parse_mode=ParseMode.HTML,
            reply_markup=admin_menu_keyboard(),
        )
        return AdminStates.DURATION_INPUT

    target = int(m.group(1))
    days = int(m.group(2))
    if days <= 0 or days > 3650:
        await update.effective_message.reply_text(
            "Kesalahan Limitasi: Jangka panjang durasi hari dibatasi antara nominal 1 dan 3650.",
            parse_mode=ParseMode.HTML,
            reply_markup=admin_menu_keyboard(),
        )
        return AdminStates.DURATION_INPUT

    store = load_store()
    u = ensure_user_record(store, target)
    exp = utc_now() + timedelta(days=days)
    u["access_expires_at"] = to_utc_iso(exp)
    save_store(store)

    await update.effective_message.reply_text(
        f"✅ Pembaruan Expire Time diterapkan untuk <code>{target}</code>.\nMasa kadaluarsa presisi dijadwalkan pada:\n<code>{u['access_expires_at']}</code>",
        parse_mode=ParseMode.HTML,
        reply_markup=admin_menu_keyboard(),
    )
    return AdminStates.MENU


async def admin_block_input_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = (update.effective_message.text or "").strip()
    m = re.fullmatch(r"(block|unblock)\s+(\d+)", text, flags=re.IGNORECASE)
    if not m:
        await update.effective_message.reply_text(
            "❌ Sintaks eksekusi penalti error. Struktur: <code>block 123456789</code> / <code>unblock 123456789</code>",
            parse_mode=ParseMode.HTML,
            reply_markup=admin_menu_keyboard(),
        )
        return AdminStates.BLOCK_INPUT

    action = m.group(1).lower()
    target = int(m.group(2))

    store = load_store()
    blocked = set(store.get("blocked", []))
    if action == "block":
        blocked.add(target)
    else:
        blocked.discard(target)
    store["blocked"] = sorted(list(blocked))
    save_store(store)

    await update.effective_message.reply_text(
        f"✅ Keputusan Restriksi dikonfirmasi untuk identitas <code>{target}</code>. Aksi dilakukan: <b>{action.upper()}</b>",
        parse_mode=ParseMode.HTML,
        reply_markup=admin_menu_keyboard(),
    )
    return AdminStates.MENU


async def admin_broadcast_input_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    message = (update.effective_message.text or "").strip()
    if not message:
        await update.effective_message.reply_text(
            "Kesalahan Pengiriman: Konten Broadcast ditolak karena kosong.",
            parse_mode=ParseMode.HTML,
            reply_markup=admin_menu_keyboard(),
        )
        return AdminStates.BROADCAST_INPUT

    store = load_store()
    users = [int(uid) for uid in store.get("users", {}).keys() if uid.isdigit()]
    success = 0
    failed = 0

    for uid in users:
        try:
            await context.bot.send_message(
                chat_id=uid,
                text=f"📢 <b>Kawat Resmi Administrasi</b>\n\n{message}",
                parse_mode=ParseMode.HTML,
                reply_markup=main_menu_keyboard(is_admin=user_is_admin(uid)),
            )
            success += 1
        except Exception:
            failed += 1

    store["stats"]["last_broadcast_at"] = to_utc_iso(utc_now())
    save_store(store)

    await update.effective_message.reply_text(
        f"✅ Rutin Pengiriman Broadcast Massa Selesai Secara Global.\n\n"
        f"Laporan Pengiriman:\n"
        f"📩 Masuk Berhasil: <b>{success}</b> transmisi\n"
        f"📉 Gagal Target: <b>{failed}</b> (Blokir bot dsb.)",
        parse_mode=ParseMode.HTML,
        reply_markup=admin_menu_keyboard(),
    )
    return AdminStates.MENU


# -----------------------------------------------------------------------------
# Global Router
# -----------------------------------------------------------------------------

async def global_text_router(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    text = (update.effective_message.text or "").strip()
    uid = current_user_id(update)

    # Hard reset controls
    if text in {MAIN_MENU_START, "Start", "/start", SUBMENU_CANCEL, "Batal"}:
        await start_handler(update, context)
        return

    if text == MAIN_MENU_CREATE_DOC:
        await menu_create_doc_handler(update, context)
        return

    if text == MAIN_MENU_HELP:
        await help_handler(update, context)
        return

    if text == MAIN_MENU_INVENTORY:
        await inventori.inventory_menu_handler(update, context)
        return

    if text == SUBMENU_BACK:
        await update.effective_message.reply_text(
            "🔙 Menuju Tampilan Menu Utama.",
            parse_mode=ParseMode.HTML,
            reply_markup=main_menu_keyboard(is_admin=user_is_admin(uid)),
        )
        return

    if text == SUBMENU_MANUAL:
        await manual_start_handler(update, context)
        return

    if text == SUBMENU_INSTANT:
        await instant_start_handler(update, context)
        return

    if text == MAIN_MENU_ADMIN:
        await admin_entry_handler(update, context)
        return

    await update.effective_message.reply_text(
        "🤖 <i>Perintah teks tidak cocok dengan navigasi antarmuka apapun. Silakan manfaatkan papan tombol interaktif di bawah.</i>",
        parse_mode=ParseMode.HTML,
        reply_markup=main_menu_keyboard(is_admin=user_is_admin(uid)),
    )


async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    logger.exception("Unhandled exception: %s", context.error)


# -----------------------------------------------------------------------------
# App Setup
# -----------------------------------------------------------------------------

def build_application() -> Application:
    token = os.getenv("TELEGRAM_TOKEN")
    if not token:
        raise RuntimeError(
            "TELEGRAM_TOKEN tidak ditemukan. Isi TELEGRAM_TOKEN di environment/.env."
        )

    app = Application.builder().token(token).build()

    inventori.register_inventory_handlers(app, guard_access)

    # Manual conversation
    manual_conv = ConversationHandler(
        entry_points=[
            CommandHandler("manual", manual_start_handler),
            MessageHandler(filters.Regex(f"^{re.escape(SUBMENU_MANUAL)}$"), manual_start_handler),
        ],
        states={
            States.ASK_UID: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_uid_handler)],
            States.ASK_PASSWORD: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_password_handler)],
            States.ASK_COOKIE: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_cookie_handler)],
            States.ASK_FILENAME_MANUAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_filename_manual_handler)],
        },
        fallbacks=[
            CommandHandler("cancel", cancel_handler),
            MessageHandler(filters.Regex(f"^{re.escape(SUBMENU_CANCEL)}$"), cancel_handler),
            CommandHandler("start", start_handler),
            MessageHandler(filters.Regex(f"^{re.escape(MAIN_MENU_START)}$"), start_handler),
            CallbackQueryHandler(cancel_callback, pattern="^cancel_input$"),
        ],
        allow_reentry=True,
        name="manual_conversation",
        persistent=False,
    )

    # Instant conversation
    instant_conv = ConversationHandler(
        entry_points=[
            CommandHandler("instan", instant_start_handler),
            MessageHandler(filters.Regex(f"^{re.escape(SUBMENU_INSTANT)}$"), instant_start_handler),
        ],
        states={
            States.ASK_INSTANT_PAYLOAD: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_instant_payload_handler)],
            States.ASK_FILENAME_INSTANT: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_filename_instant_handler)],
        },
        fallbacks=[
            CommandHandler("cancel", cancel_handler),
            MessageHandler(filters.Regex(f"^{re.escape(SUBMENU_CANCEL)}$"), cancel_handler),
            CommandHandler("start", start_handler),
            MessageHandler(filters.Regex(f"^{re.escape(MAIN_MENU_START)}$"), start_handler),
            CallbackQueryHandler(cancel_callback, pattern="^cancel_input$"),
        ],
        allow_reentry=True,
        name="instant_conversation",
        persistent=False,
    )

    # Admin conversation
    admin_conv = ConversationHandler(
        entry_points=[
            CommandHandler("admin", admin_entry_handler),
            MessageHandler(filters.Regex(f"^{re.escape(MAIN_MENU_ADMIN)}$"), admin_entry_handler),
        ],
        states={
            AdminStates.MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, admin_menu_router)],
            AdminStates.WHITELIST_INPUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, admin_whitelist_input_handler)],
            AdminStates.DURATION_INPUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, admin_duration_input_handler)],
            AdminStates.BLOCK_INPUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, admin_block_input_handler)],
            AdminStates.BROADCAST_INPUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, admin_broadcast_input_handler)],
        },
        fallbacks=[
            CommandHandler("cancel", cancel_handler),
            MessageHandler(filters.Regex(f"^{re.escape(SUBMENU_CANCEL)}$"), cancel_handler),
            CommandHandler("start", start_handler),
            MessageHandler(filters.Regex(f"^{re.escape(MAIN_MENU_START)}$"), start_handler),
        ],
        allow_reentry=True,
        name="admin_conversation",
        persistent=False,
    )

    app.add_handler(CommandHandler("start", start_handler))
    app.add_handler(CommandHandler("help", help_handler))
    app.add_handler(CommandHandler("buat", menu_create_doc_handler))
    app.add_handler(manual_conv)
    app.add_handler(instant_conv)
    app.add_handler(admin_conv)

    # Global text router (outside active conversation)
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, global_text_router))

    app.add_error_handler(error_handler)
    return app


def main() -> None:
    try:
        app = build_application()
        logger.info("FBDocBot is now actively polling for updates...")
        app.run_polling(drop_pending_updates=True)
    except Exception:
        logger.exception("Terdeteksi fatal error saat inisialisasi bot system.")
        raise


if __name__ == "__main__":
    main()
